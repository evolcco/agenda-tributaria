from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from holidays import adjust_due_date

SCHEMA_VERSION = "1.1.0"

PT_MONTHS = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

PT_MONTH_LABELS = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def _agenda_period_from_url(page_url: str | None) -> tuple[int, int] | None:
    if not page_url:
        return None
    match = re.search(r"/(20\d{2})/([A-Za-zÀ-ÿ-]+)(?:/|$)", page_url)
    if not match:
        return None
    year = int(match.group(1))
    slug = unicodedata.normalize("NFD", match.group(2).lower())
    slug = "".join(ch for ch in slug if unicodedata.category(ch) != "Mn")
    token = slug.replace("-", " ").split(" ")[0]
    month = PT_MONTHS.get(token)
    return (year, month) if month else None


def _build_due_dates(
    due_day: int | None,
    agenda_year: int | None,
    agenda_month: int | None,
) -> dict[str, Any]:
    """Calcula due_date e effective_due_date.

    Retorna campos prontos para mesclar no item. Quando não há base suficiente
    (due_day inválido ou ano/mês desconhecido), retorna campos nulos.
    """
    if (
        due_day is None
        or due_day <= 0
        or agenda_year is None
        or agenda_month is None
    ):
        return {
            "due_date": None,
            "effective_due_date": None,
            "adjusted": False,
            "adjustment_reason": None,
        }

    try:
        original = date(agenda_year, agenda_month, due_day)
    except ValueError:
        return {
            "due_date": None,
            "effective_due_date": None,
            "adjusted": False,
            "adjustment_reason": None,
        }

    adj = adjust_due_date(original)
    return {
        "due_date": adj.original.isoformat(),
        "effective_due_date": adj.effective.isoformat(),
        "adjusted": adj.adjusted,
        "adjustment_reason": adj.reason,
    }


@dataclass(frozen=True)
class Competence:
    year: int
    month: int

    @property
    def label(self) -> str:
        names = {
            1: "Janeiro",
            2: "Fevereiro",
            3: "Março",
            4: "Abril",
            5: "Maio",
            6: "Junho",
            7: "Julho",
            8: "Agosto",
            9: "Setembro",
            10: "Outubro",
            11: "Novembro",
            12: "Dezembro",
        }
        return f"{names[self.month]}/{self.year}"


def _norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _header_key(value: Any) -> str:
    text = _norm(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.lower()


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = _norm(value)
    if not text:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except ValueError:
        return None


def _find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    required_headers: Iterable[str],
) -> tuple[int, dict[str, int]]:
    required = {_header_key(h): h for h in required_headers}

    for row in range(1, min(ws.max_row, 30) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value:
                headers[_header_key(value)] = col

        if all(key in headers for key in required):
            return row, headers

    missing = ", ".join(required_headers)
    raise ValueError(f"Nao foi possivel localizar cabecalho com: {missing}")


def _iter_data_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    max_blank_streak: int = 3,
) -> Iterator[tuple[int, list[Any]]]:
    blank_streak = 0
    for row in range(start_row, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if all(v is None or _norm(v) == "" for v in values):
            blank_streak += 1
            if blank_streak >= max_blank_streak:
                break
            continue

        blank_streak = 0
        yield row, values


def _parse_tributos(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    agenda_year: int | None,
    agenda_month: int | None,
) -> list[dict[str, Any]]:
    required = [
        "Dia de vencimento",
        "Código de Receita",
        "Grupo de Tributo",
        "Descrição",
        "Período de Apuração",
        "Periodicidade",
        "Documento de Arrecadação",
    ]
    header_row, header = _find_header_row(ws, required)

    result: list[dict[str, Any]] = []
    idx = 1
    for row, _values in _iter_data_rows(ws, header_row + 1):
        due_day = _to_int(ws.cell(row, header[_header_key("Dia de vencimento")]).value)
        code = _to_int(ws.cell(row, header[_header_key("Código de Receita")]).value)
        group = _norm(ws.cell(row, header[_header_key("Grupo de Tributo")]).value)
        description = _norm(ws.cell(row, header[_header_key("Descrição")]).value)
        apuration_period = _norm(ws.cell(row, header[_header_key("Período de Apuração")]).value)
        periodicity = _norm(ws.cell(row, header[_header_key("Periodicidade")]).value)
        payment_document = _norm(
            ws.cell(row, header[_header_key("Documento de Arrecadação")]).value
        )

        if not any([due_day is not None, code, group, description]):
            continue

        due_type = "daily" if (due_day is None or due_day <= 0) else "fixed_day"
        item = {
            "id": f"T-{idx:04d}",
            "due_day": due_day,
            "due_type": due_type,
            **_build_due_dates(due_day, agenda_year, agenda_month),
            "code": code,
            "group": group,
            "description": description,
            "apuration_period": apuration_period,
            "periodicity": periodicity,
            "payment_document": payment_document,
        }
        result.append(item)
        idx += 1

    return result


def _parse_declaracoes(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    agenda_year: int | None,
    agenda_month: int | None,
) -> list[dict[str, Any]]:
    required = [
        "Prazo de Apresentação",
        "Interessado",
        "Declarações, Demonstrativos e Documentos",
        "Período de Referência",
        "Base Normativa",
    ]
    header_row, header = _find_header_row(ws, required)

    result: list[dict[str, Any]] = []
    idx = 1
    for row, _ in _iter_data_rows(ws, header_row + 1):
        due_day = _to_int(ws.cell(row, header[_header_key("Prazo de Apresentação")]).value)
        interested = _norm(ws.cell(row, header[_header_key("Interessado")]).value)
        description = _norm(
            ws.cell(
                row,
                header[_header_key("Declarações, Demonstrativos e Documentos")],
            ).value
        )
        reference_period = _norm(ws.cell(row, header[_header_key("Período de Referência")]).value)
        legal_basis = _norm(ws.cell(row, header[_header_key("Base Normativa")]).value)

        if not any([due_day is not None, interested, description]):
            continue

        item = {
            "id": f"D-{idx:04d}",
            "due_day": due_day,
            **_build_due_dates(due_day, agenda_year, agenda_month),
            "interested": interested,
            "description": description,
            "reference_period": reference_period,
            "legal_basis": legal_basis,
        }
        result.append(item)
        idx += 1

    return result


def _infer_competence(declaracoes: list[dict[str, Any]]) -> Competence | None:
    candidates: list[tuple[int, int]] = []

    for item in declaracoes:
        ref = item.get("reference_period", "")
        match = re.search(r"([A-Za-zçÇãÃéÉ]+)\s*/\s*(\d{4})", ref)
        if not match:
            continue
        month_name = match.group(1).lower().replace("ç", "c")
        year = int(match.group(2))
        month = PT_MONTHS.get(month_name)
        if month is not None:
            candidates.append((year, month))

    if not candidates:
        return None

    year, month = Counter(candidates).most_common(1)[0][0]
    return Competence(year=year, month=month)


def parse_xlsx(
    input_path: Path,
    *,
    competence_year: int | None = None,
    competence_month: int | None = None,
    agenda_year: int | None = None,
    agenda_month: int | None = None,
    monthly_page_url: str | None = None,
    xlsx_url: str | None = None,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if "Tributos" not in wb.sheetnames or "Declarações" not in wb.sheetnames:
        raise ValueError("Planilha nao contem abas esperadas: 'Tributos' e 'Declarações'.")

    if agenda_year is None or agenda_month is None:
        period = _agenda_period_from_url(monthly_page_url)
        if period is not None:
            agenda_year = agenda_year or period[0]
            agenda_month = agenda_month or period[1]

    tributos = _parse_tributos(wb["Tributos"], agenda_year, agenda_month)
    declaracoes = _parse_declaracoes(wb["Declarações"], agenda_year, agenda_month)

    inferred = _infer_competence(declaracoes)
    final_year = competence_year or (inferred.year if inferred else None)
    final_month = competence_month or (inferred.month if inferred else None)

    competence = None
    if final_year and final_month:
        competence = Competence(year=final_year, month=final_month)

    agenda = None
    if agenda_year and agenda_month:
        agenda = {
            "year": agenda_year,
            "month": agenda_month,
            "label": f"{PT_MONTH_LABELS[agenda_month]}/{agenda_year}",
        }

    adjusted_count = sum(
        1 for item in (*tributos, *declaracoes) if item.get("adjusted")
    )

    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.now(UTC).isoformat(),
        "agenda": agenda,
        "competence": {
            "year": competence.year if competence else None,
            "month": competence.month if competence else None,
            "label": competence.label if competence else None,
        },
        "source": {
            "monthly_page_url": monthly_page_url,
            "xlsx_url": xlsx_url,
            "xlsx_file_name": input_path.name,
        },
        "counts": {
            "tributos": len(tributos),
            "declaracoes": len(declaracoes),
            "adjusted": adjusted_count,
        },
        "tributos": tributos,
        "declaracoes": declaracoes,
    }


def write_json(output_path: Path, payload: dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Converte XLSX da agenda tributaria em JSON")
    parser.add_argument("--input", required=True, type=Path, help="Caminho para o arquivo .xlsx")
    parser.add_argument("--output", required=True, type=Path, help="Caminho de saida .json")
    parser.add_argument("--year", type=int, help="Ano da competencia (apuracao)")
    parser.add_argument("--month", type=int, help="Mes da competencia (1-12)")
    parser.add_argument("--agenda-year", type=int, help="Ano da agenda (mes dos vencimentos)")
    parser.add_argument("--agenda-month", type=int, help="Mes da agenda (1-12)")
    parser.add_argument("--monthly-page-url", help="URL da pagina mensal da agenda")
    parser.add_argument("--xlsx-url", help="URL original da planilha")
    return parser


def main() -> None:
    args = _build_cli().parse_args()
    payload = parse_xlsx(
        args.input,
        competence_year=args.year,
        competence_month=args.month,
        agenda_year=args.agenda_year,
        agenda_month=args.agenda_month,
        monthly_page_url=args.monthly_page_url,
        xlsx_url=args.xlsx_url,
    )
    write_json(args.output, payload)


if __name__ == "__main__":
    main()
